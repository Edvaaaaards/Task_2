from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0

for row in range(2,ws.max_row+1):
    hours=ws["B" + str(row)].value
    rate=ws["C" + str(row)].value
    if (type(hours)!=str and type(rate)!=str):
        
        salary=hours*rate
        ws["D"+str(row)].value=salary        
        if (salary>3000):
            total=total+1
            

print(total)
wb.close()
